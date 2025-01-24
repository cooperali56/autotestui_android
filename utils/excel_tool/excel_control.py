import os

import openpyxl
import xlrd as xlrd

from common.setting import ensure_path_sep, find_file_path
from utils.common_tool.common_data import handle_excel_dict


def get_all_excel(file_name, sheet) -> list:
    """
    获取所有Excel文件
    :return:
    """
    return ExcelControl(file_name=file_name).get_excel_data(sheet_name=sheet)


class ExcelControl:
    """
    处理Excel
    """

    def __init__(self, file_name):
        self.workbook = None
        self.sheet_data = None
        self.file_path = None
        self.file_name = file_name

    def _get_file_path(self):
        if os.path.sep in self.file_name:
            return ensure_path_sep(f"\\{self.file_name}")
        else:
            return find_file_path(self.file_name)

    def get_excel_data(self, sheet_name) -> list:
        excel_data = []
        self.file_path = self._get_file_path()

        if self.file_name.endswith(".xlsx"):
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.sheet_data = self.workbook[sheet_name]
            excel_data = [row for row in self.sheet_data.iter_rows(values_only=True)]
            self.workbook.close()
        elif self.file_name.endswith(".xls"):
            self.workbook = xlrd.open_workbook(self.file_path)
            self.sheet_data = self.workbook.sheet_by_name(sheet_name)
            excel_data = [list(self.sheet_data.row_values(row)) for row in range(self.sheet_data.nrows)]
            self.workbook.release_resources()
        excel_data = handle_excel_dict(excel_data)
        return excel_data

    def cell_input(self, case_id, actual, input_data) -> None:
        """
        写入Excel数据
        :param case_id:     case-id 用于定位单元格
        :param actual:    actual 用于定位单元格
        :param input_data:  需要写入的数据
        :return:
        """
        # 获取需要输入case-id的单元格坐标
        row_int = None
        column_int = None
        for row in self.sheet_data.iter_rows():
            for cell in row:
                if cell.value == case_id:
                    row_int = cell.row
                    break
                if cell.value == actual:
                    column_int = cell.column
                    break
        # 根据后缀选择写入方式和保存方式
        self.sheet_data.cell(row=row_int, column=column_int, value=input_data)
        self.workbook.save(self.file_path)
        if self.file_name.endswith(".xlsx"):
            self.workbook.close()
        elif self.file_name.endswith(".xls"):
            self.workbook.release_resources()
